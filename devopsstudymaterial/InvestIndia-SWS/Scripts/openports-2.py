import boto3
import openpyxl

region_name = 'ap-south-1'
ec2_client = boto3.client('ec2',region_name=region_name)

security_groups = ec2_client.describe_security_groups()['SecurityGroups']

workbook = openpyxl.Workbook()

inbound_sheet = workbook.create_sheet(title='Inbound Rules', index=0)
outbound_sheet = workbook.create_sheet(title='Outbound Rules', index=1)

headers = ['Security Group Name', 'SG Description', 'Security Group ID', 'Region', 'VPC ID', 'Ports Open for 0.0.0.0/0', 'Rule Description', 'Rule Type', 'Resolution']
inbound_sheet.append(headers)
outbound_sheet.append(headers)

for group in security_groups:
    group_id = group['GroupId']
    group_name = group['GroupName']
    region = ec2_client.meta.region_name
    vpc_id = group['VpcId']
    sg_description = group['Description']
    
    for rule in group['IpPermissions']:
        for ip_range in rule['IpRanges']:
            if ip_range['CidrIp'] == '0.0.0.0/0':
                from_port = rule.get('FromPort', 'All')
                to_port = rule.get('ToPort', 'All')
                port_range = f"{from_port}-{to_port}"
                rule_description = ip_range.get('Description', 'No Description')
                rule_type = 'Inbound'
                inbound_sheet.append([group_name, sg_description, group_id, region, vpc_id, port_range,rule_description, rule_type])
    
    for rule in group['IpPermissionsEgress']:
        for ip_range in rule['IpRanges']:
            if ip_range['CidrIp'] == '0.0.0.0/0':
                from_port = rule.get('FromPort', 'All')
                to_port = rule.get('ToPort', 'All')
                port_range = f"{from_port}-{to_port}"
                rule_description = ip_range.get('Description', 'No Description')
                rule_type = 'Outbound'
                outbound_sheet.append([group_name, sg_description, group_id, region, vpc_id, port_range,rule_description, rule_type])

workbook.save('open_ports_report.xlsx')
print("Excel file 'open_ports_report.xlsx' saved successfully.")

